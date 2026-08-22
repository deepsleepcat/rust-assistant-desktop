r"""
从用户提供的 xlsx 代码表审计并增量补全 public/data。

默认只输出审计结果；使用 --apply 才会写入 code.json/translations.json。

用法：
  python scripts/merge-xlsx-data.py "W:\mao\tx\代码表1.15超强版(兼容)0830.xlsx"
  python scripts/merge-xlsx-data.py "W:\mao\tx\代码表1.15超强版(兼容)0830.xlsx" --apply

来源表「代码表HX」：
  - 节行：B=中文节名，D=[英文节名]
  - 字段行：B=英文代码（可带或不带末尾冒号），C=中文翻译，D=描述，E=例子，F=值类型
"""
import json
import re
import sys
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
CODE_JSON = ROOT / "public" / "data" / "code.json"
TRANS_JSON = ROOT / "public" / "data" / "translations.json"
SHEET_NAME = "代码表HX"
SECTION_RE = re.compile(r"^\[([A-Za-z_][A-Za-z0-9_]*)\]$")
# 支持工作簿中的 builtFrom_#_*、displayText_{LANG}、arm#_[time] 等模板键。
FIELD_RE = re.compile(r"^([A-Za-z_][A-Za-z0-9_#{}\[\].-]*):?$")


@dataclass(frozen=True)
class Row:
    row: int
    kind: str
    key: str
    translate: str
    description: str
    demo: str
    value_type: str
    section: str


def _under_root(path: Path) -> Path:
    """解析并校验写入路径必须在项目根内。"""
    resolved = path.resolve()
    root = ROOT.resolve()
    if resolved != root and root not in resolved.parents:
        raise SystemExit(f"路径越出项目根目录: {path}")
    return resolved


def text(value: object) -> str:
    return str(value).strip() if value is not None else ""


def normalized_key(key: str) -> str:
    """将不同来源的数字模板写法归一，便于识别 # 与 {NUM} 的同义字段。"""
    return key.lower().replace("{num}", "#")


def extract_rows(xlsx_path: str) -> list[Row]:
    workbook = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=False, keep_links=False)
    if SHEET_NAME not in workbook.sheetnames:
        raise SystemExit(f"工作簿缺少工作表: {SHEET_NAME}")
    sheet = workbook[SHEET_NAME]
    rows: list[Row] = []
    current_section = ""
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        code = text(row[1] if len(row) > 1 else None)
        translate = text(row[2] if len(row) > 2 else None)
        description = text(row[3] if len(row) > 3 else None)
        demo = text(row[4] if len(row) > 4 else None)
        value_type = text(row[5] if len(row) > 5 else None)

        section_match = SECTION_RE.fullmatch(description)
        if section_match:
            current_section = section_match.group(1)
            if code and code != "代码":
                rows.append(Row(row_number, "section", current_section, code, "", "", "", current_section))
            continue

        field_match = FIELD_RE.fullmatch(code)
        if field_match:
            rows.append(
                Row(
                    row_number,
                    "code",
                    field_match.group(1),
                    translate,
                    description,
                    demo,
                    value_type,
                    current_section,
                )
            )
    return rows


def report_conflicts(rows: list[Row]) -> None:
    grouped: dict[tuple[str, str], list[Row]] = defaultdict(list)
    for row in rows:
        if row.kind == "code":
            grouped[(normalized_key(row.key), row.section.lower())].append(row)
    duplicates = {key: items for key, items in grouped.items() if len(items) > 1}
    conflicts = {
        key: items
        for key, items in duplicates.items()
        if len({(item.translate, item.value_type) for item in items}) > 1
    }
    print(f"解析字段 {sum(row.kind == 'code' for row in rows)} 条，节 {sum(row.kind == 'section' for row in rows)} 条")
    print(f"重复键 {len(duplicates)} 组，字段元数据冲突 {len(conflicts)} 组")
    for (key, section), items in list(duplicates.items())[:20]:
        details = "; ".join(f"第{item.row}行「{item.translate}」/{item.value_type or '无类型'}" for item in items)
        marker = " [元数据冲突]" if (key, section) in conflicts else ""
        print(f"  {section or '<全局>'}.{key}: {details}{marker}")
    if len(duplicates) > 20:
        print(f"  … 其余 {len(duplicates) - 20} 组未展开")


def main() -> None:
    args = [arg for arg in sys.argv[1:] if arg != "--apply"]
    apply = "--apply" in sys.argv[1:]
    xlsx = args[0] if args else r"W:\mao\tx\代码表1.15超强版(兼容)0830.xlsx"
    rows = extract_rows(xlsx)
    report_conflicts(rows)

    with _under_root(CODE_JSON).open(encoding="utf-8") as file:
        codes = json.load(file)
    with _under_root(TRANS_JSON).open(encoding="utf-8") as file:
        translations = json.load(file)

    existing_codes = {normalized_key(text(item.get("code"))) for item in codes.get("data", [])}
    existing_translations = {(text(item.get("en")), text(item.get("zh"))) for item in translations.get("words", [])}
    additions = [row for row in rows if row.kind == "code" and normalized_key(row.key) not in existing_codes]
    missing_translation = [row for row in additions if not row.translate]
    print(f"代码表缺失字段 {len(additions)} 条，缺中文译名 {len(missing_translation)} 条")
    for row in additions[:30]:
        print(f"  第{row.row}行 {row.key}: {row.translate or '<缺译名>'}")
    if len(additions) > 30:
        print(f"  … 其余 {len(additions) - 30} 条未展开")

    if not apply:
        print("审计完成；未写入数据。使用 --apply 才会追加有中文译名的缺失字段。")
        return

    new_codes = []
    new_translations = []
    for row in additions:
        if not row.translate:
            continue
        new_codes.append(
            {
                "code": row.key,
                "translate": row.translate,
                "description": row.description or row.translate,
                "type": row.value_type or "unknown",
                "addVersion": 0,
                "removeVersion": -1,
                "section": row.section,
                "demo": row.demo,
            }
        )
        pair = (row.key, row.translate)
        if pair not in existing_translations:
            new_translations.append({"en": row.key, "zh": row.translate})
            existing_translations.add(pair)

    codes["data"].extend(new_codes)
    translations["words"].extend(new_translations)
    _under_root(CODE_JSON).write_text(json.dumps(codes, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    _under_root(TRANS_JSON).write_text(json.dumps(translations, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"已追加字段 {len(new_codes)} 条，翻译对 {len(new_translations)} 条")


if __name__ == "__main__":
    main()
